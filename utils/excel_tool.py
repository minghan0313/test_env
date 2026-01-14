import openpyxl
from openpyxl.styles import PatternFill, Font

class ExcelProcessor:
    """通用的 Excel 处理工具，支持按单元格写入和样式设置"""
    
    @staticmethod
    def load_workbook(path):
        return openpyxl.load_workbook(path)

    @staticmethod
    def set_cell_value(sheet, cell_coord, value, is_alert=False):
        """写入数据，如果 is_alert 为 True，则将单元格标红"""
        cell = sheet[cell_coord]
        cell.value = value
        
        if is_alert:
            # 红色填充，白色文字（展示警示效果）
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            
    @staticmethod
    def save(wb, path):
        wb.save(path)